import csv, os
import openpyxl

"""
    Converts csv files in WoL (Web of Life) format to the chosen IWDB format.
    this means breaking species names to the {genus} {species} format, and 
    transposing the network(*1) from rows=plants and columns=pollinators
    to rows=pollinators and columns=plants.
    
    (*1) Transposition assumes plants are rows in all WoL files according
    to cursory check.
"""

WOL_DIR = "C:\\Personal\\University\\Lab\\WoL"
CONVERT_DIR = "C:\\Personal\\University\\Lab\\WoL\\converted"
REFERENCES = dict()


def setup_workbook(network_name):
    wb = openpyxl.Workbook()
    sheet = wb.active

    sheet['A1'] = str(REFERENCES[network_name])
    sheet['C1'] = "plant_ge"
    sheet['C2'] = "plant_sp"
    sheet['C3'] = "no."
    sheet['A3'] = "pol_ge"
    sheet['B3'] = "pol_sp"
    return wb


def extract_plant(plant_readings, sheet, column_index):
    row_index = 4
    try:
        plant_name = next(plant_readings)
    except StopIteration:
        print("stopped, no plant name given in: " + plant_readings)
        return

    genus = plant_name.split(" ")[0]
    species = plant_name[len(genus) + 1:]

    sheet.cell(row=1, column=column_index, value=genus)
    sheet.cell(row=2, column=column_index, value=species)
    sheet.cell(row=3, column=column_index, value=column_index)
    for read in plant_readings:
        sheet.cell(row=row_index, column=column_index, value=read)
        row_index += 1


def extract_pollinators(pollinator_list, sheet):
    row_index = 4
    for pollinator in pollinator_list:
        genus = pollinator.split(" ")[0]
        species = pollinator[len(genus) + 1:]

        sheet.cell(row=row_index, column=1, value=genus)
        sheet.cell(row=row_index, column=2, value=species)
        sheet.cell(row=row_index, column=3, value=row_index)
        row_index += 1


def convert_wol_excel(dir, filename):
    target_book = setup_workbook(filename.split(".")[0])
    sheet = target_book.active
    column_index = 4

    with open(dir + "\\" + filename) as wol_file:
        wol_reader = csv.reader(wol_file)
        extract_pollinators(next(wol_reader)[1:], sheet)

        for plant in wol_reader:
            extract_plant(iter(plant), sheet, column_index)
            column_index += 1
    target_book.save(filename=CONVERT_DIR + "\\" + filename.split(".")[0] + ".xlsx")
    print("converted file created " + filename.split(".")[0])


def generate_references():
    with open(WOL_DIR + "\\references.csv") as reference_file:
        ref_reader = csv.reader(reference_file)
        ref_params = next(ref_reader)

        for row in ref_reader:
            REFERENCES[row[0]] = {ref_params[i]: row[i] for i in range(1, len(row))}
    print("generated file references for all networks")


if __name__ == "__main__":
    generate_references()
    for filename in os.listdir(WOL_DIR):
        if filename.endswith(".csv") and filename != "references.csv":
            convert_wol_excel(WOL_DIR, filename)
