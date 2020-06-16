import os
import openpyxl
import pandas as pd

"""
    Checks for duplicate networks in a two steps method:
    1. Hash all networks by their excel dimensions. This provides a quick method to
       find "suspicious" groups of networks with the same dimensions, those will go to
       the second step.
    2. Hash all networks with the same dimensions by a sorted list of all of their interactions,
       where two networks will be considered duplicate if they have the same interactions list.
    
    Assumptions:
    1. Metadata might be missing and cannot be trusted.
    2. A network is captured fully by all of it's interactions 
        (that is, no isolated nodes)
    3. Solves both duplicates inside DBs, and among DBs. 
"""

TEST_BASE = "C:\\Personal\\University\\Lab"
TEST_LIST = ["WoL\\converted"]#, "Mangal", "IWDB"]


def get_excel_dimensions(filename):
    wb = openpyxl.load_workbook(filename)
    return wb.active.max_row, wb.active.max_column


def iterate_dirlist(dirlist, filetype=""):
    for directory in dirlist:
        for filename in os.listdir(directory):
            if filename.endswith(filetype):
                yield directory + "\\" + filename


def hash_by_dimension(dirlist):
    dimension_dict = {}

    for filename in iterate_dirlist(dirlist, filetype="xlsx"):
        file_dims = get_excel_dimensions(filename)
        if file_dims in dimension_dict:
            dimension_dict[file_dims].append(filename)
        else:
            dimension_dict[file_dims] = [filename]
    return dimension_dict


def hash_by_interaction(file_list):
    raise NotImplementedError


if __name__ == "__main__":
    dim_hash = hash_by_dimension([TEST_BASE + "\\" + e for e in TEST_LIST])
    possible_duplicates = []

    for dim, file_list in dim_hash.items():
        if len(file_list) > 1:
            possible_duplicates.append(hash_by_interaction(file_list))
